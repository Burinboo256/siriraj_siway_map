#!/usr/bin/env python3
"""Convert siriraj_data_location.xlsx to data/siriraj_data_location.json."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from urllib.parse import quote

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "siriraj_data_location.xlsx"
OUTPUT_JSON = ROOT / "data" / "siriraj_data_location.json"
BUILDING_CATEGORIES = {"building", "dormitory"}


def to_str(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0") and s.replace(".0", "", 1).replace("-", "").isdigit():
        s = s[:-2]
    return s


def parse_lat_lon(raw: Any) -> tuple[float | None, float | None]:
    if raw is None:
        return None, None
    txt = str(raw).strip()
    if "," not in txt:
        return None, None
    a, b = [x.strip() for x in txt.split(",", 1)]
    try:
        return float(a), float(b)
    except ValueError:
        return None, None


def icon_for(group: str, category: str) -> str:
    by_group = {
        "clinical": "🩺",
        "academic": "🎓",
        "faculty": "🏫",
        "service": "🛠️",
        "admin": "🏛️",
        "hospital": "🏥",
        "research": "🔬",
        "memorial": "🕊️",
        "amenity": "🛎️",
        "parking": "🅿️",
        "student": "🎒",
        "resident": "👩‍⚕️",
        "nursing": "👩‍⚕️",
        "museum": "🏛️",
    }
    by_category = {
        "building": "🏢",
        "facility": "🧩",
        "museum": "🏛️",
        "landmark": "📍",
        "dormitory": "🏠",
    }
    return by_group.get(group) or by_category.get(category) or "📍"


def google_maps_url(lat: float | None, lon: float | None) -> str:
    if lat is None or lon is None:
        return ""
    destination = quote(f"{lat},{lon}", safe="")
    return f"https://www.google.com/maps/dir/?api=1&destination={destination}&travelmode=walking"


def entity_type_for(category: str) -> str:
    return "building" if category in BUILDING_CATEGORIES else "unit"


def main() -> None:
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    raw_rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 13)]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        raw_rows.append(vals)

    parsed = []
    for vals in raw_rows:
        raw_id, map_no, name_th, name_en, category, group, facilities, latmix, _lon, desc_th, _desc_en, _img = vals
        lat, lon = parse_lat_lon(latmix)

        fac = []
        if facilities:
            fac = [x.strip() for x in str(facilities).split("|") if x and str(x).strip()]

        parsed.append(
            {
                "id": to_str(raw_id),
                "map_no": to_str(map_no),
                "name_th": to_str(name_th),
                "name_en": to_str(name_en),
                "category": to_str(category),
                "group": to_str(group),
                "facilities": fac,
                "lat": lat,
                "lon": lon,
                "description_th": to_str(desc_th),
            }
        )

    coords = [p for p in parsed if p["lat"] is not None and p["lon"] is not None]
    min_lat = min((p["lat"] for p in coords), default=0.0)
    max_lat = max((p["lat"] for p in coords), default=1.0)
    min_lon = min((p["lon"] for p in coords), default=0.0)
    max_lon = max((p["lon"] for p in coords), default=1.0)
    lat_span = (max_lat - min_lat) or 1.0
    lon_span = (max_lon - min_lon) or 1.0

    def to_xy(lat: float, lon: float) -> tuple[float, float]:
        x = 8 + ((lon - min_lon) / lon_span) * 84
        y = 10 + ((max_lat - lat) / lat_span) * 80
        return round(x, 1), round(y, 1)

    places = []
    for p in parsed:
        map_x = map_y = None
        if p["lat"] is not None and p["lon"] is not None:
            map_x, map_y = to_xy(p["lat"], p["lon"])
        entity_type = entity_type_for(p["category"])
        floor_label = f"ตำแหน่ง {p['map_no']}" if p["map_no"] else "-"

        aliases = [p["name_th"], p["name_en"], p["id"], p["map_no"], p["category"], p["group"], *p["facilities"]]
        dedup = []
        seen = set()
        for a in aliases:
            a = (a or "").strip()
            if not a:
                continue
            k = a.lower()
            if k in seen:
                continue
            seen.add(k)
            dedup.append(a)

        places.append(
            {
                "id": p["id"],
                "name_th": p["name_th"] or p["id"],
                "name_en": p["name_en"] or p["name_th"] or p["id"],
                "aliases": dedup,
                "entity_type": entity_type,
                "unit_type": p["category"] or "-",
                "building_id": p["id"] if entity_type == "building" else "",
                "building_name_th": p["name_th"] if entity_type == "building" else "",
                "building_name_en": p["name_en"] if entity_type == "building" else "",
                "category": p["category"] or "-",
                "building": p["group"] or "-",
                "floor": floor_label,
                "floor_label": floor_label,
                "room": "",
                "description": p["description_th"] or (f"สถานที่กลุ่ม {p['group']}" if p["group"] else "ข้อมูลจากไฟล์ Excel"),
                "phone": "-",
                "opening_hours": "-",
                "nearby_landmark": ", ".join(p["facilities"]) if p["facilities"] else "-",
                "map_x": map_x,
                "map_y": map_y,
                "is_accessible": True,
                "icon": icon_for(p["group"], p["category"]),
                "lat": p["lat"],
                "lon": p["lon"],
                "google_maps_url": google_maps_url(p["lat"], p["lon"]),
            }
        )

    buildings = [
        {
            "id": pl["id"],
            "name_th": pl["name_th"],
            "name_en": pl["name_en"],
            "floors": None,
            "lat": pl["lat"],
            "lon": pl["lon"],
            "map_x": pl["map_x"],
            "map_y": pl["map_y"],
            "google_maps_url": pl["google_maps_url"],
        }
        for pl in places
        if pl["entity_type"] == "building"
    ]

    start_nodes = [
        {
            "id": pl["id"],
            "name": pl["name_th"],
            "type": pl["category"],
            "map_x": pl["map_x"],
            "map_y": pl["map_y"],
        }
        for pl in places
        if isinstance(pl["map_x"], (int, float)) and isinstance(pl["map_y"], (int, float))
    ][:4]

    payload = {
        "schemaVersion": 2,
        "buildings": buildings,
        "places": places,
        "startNodes": start_nodes,
        "routeSteps": {
            "default": [
                "เริ่มจากจุดหลักที่ใกล้ที่สุดบนแผนที่",
                "เดินตามทางหลักและป้ายอาคาร",
                "หากไม่แน่ใจให้สอบถามจุดประชาสัมพันธ์",
            ]
        },
        "quick": [
            ["OPD", "opd", "🏥"],
            ["ห้องยา", "ยา", "💊"],
            ["เจาะเลือด", "เจาะเลือด", "🧪"],
            ["ห้องน้ำ", "ห้องน้ำ", "🚻"],
            ["ลิฟต์", "ลิฟต์", "🛗"],
            ["ที่จอดรถ", "parking", "🚗"],
            ["รถรับส่ง", "shuttle", "🚌"],
            ["ถามทาง", "ประชาสัมพันธ์", "ℹ️"],
        ],
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON} with {len(places)} places, {len(start_nodes)} start nodes")


if __name__ == "__main__":
    main()
